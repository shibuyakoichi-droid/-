"""
楽天 商品別売上データ → SKU運用ランク付与スクリプト
今年：20260412-20260512_Item_SalesList.csv
去年：20250412-20250512_Item_SalesList.csv
"""

import pandas as pd
from pathlib import Path

# ---- ファイルパス ----
BASE = Path(r"c:\Users\81804\Downloads")
FILE_THIS = BASE / "20260412-20260512_Item_SalesList.csv"
FILE_LAST = BASE / "20250412-20250512_Item_SalesList.csv"
OUTPUT_XLSX = BASE / "SKU運用ランク_20260412-20260512.xlsx"
OUTPUT_CSV  = BASE / "SKU運用ランク_20260412-20260512.csv"

# ---- CSV読み込み（ヘッダーは7行目 = skiprows=6） ----
COLS = ["商品名", "商品管理番号", "商品番号", "平均単価", "売上個数", "売上", "売上件数"]

def load(path):
    df = pd.read_csv(path, skiprows=6, header=0, encoding="utf-8-sig",
                     names=COLS, usecols=range(7))
    df["売上"] = pd.to_numeric(df["売上"].astype(str).str.replace(",", ""), errors="coerce")
    df = df.dropna(subset=["商品管理番号", "売上"])
    return df[["商品管理番号", "商品名", "売上"]].copy()

df_this = load(FILE_THIS)
df_last = load(FILE_LAST)

# ---- 今年データを売上降順でソート ----
df_this = df_this.sort_values("売上", ascending=False).reset_index(drop=True)

# ---- 売上順位・売上ランク ----
df_this["売上順位"] = df_this.index + 1

def sales_rank(rank):
    if rank <= 30:   return "A"
    if rank <= 60:   return "B"
    if rank <= 100:  return "C"
    return "D"

df_this["売上ランク"] = df_this["売上順位"].apply(sales_rank)

# ---- 去年売上をマージ ----
df_last_slim = df_last[["商品管理番号", "売上"]].rename(columns={"売上": "去年売上"})
df = df_this.merge(df_last_slim, on="商品管理番号", how="left")

# ---- 昨対増減率・昨対ランク ----
def yoy_rate(row):
    if pd.isna(row["去年売上"]) or row["去年売上"] == 0:
        return None
    return (row["売上"] - row["去年売上"]) / row["去年売上"] * 100

def yoy_rank(row):
    if pd.isna(row["去年売上"]) or row["去年売上"] == 0:
        return "新商品"
    r = row["昨対増減率"]
    if r >= 0:      return "A"
    if r >= -20:    return "B"
    if r >= -50:    return "C"
    return "D"

df["昨対増減率"] = df.apply(yoy_rate, axis=1).round(1)
df["昨対ランク"] = df.apply(yoy_rank, axis=1)
df["粗利ランク"] = ""

# ---- 列整形・名称合わせ ----
df = df.rename(columns={"売上": "今年売上", "商品管理番号": "品番"})
FINAL_COLS = ["品番", "商品名", "今年売上", "去年売上", "売上順位", "売上ランク",
              "昨対増減率", "昨対ランク", "粗利ランク"]
df = df[FINAL_COLS]

# ---- Excel出力（見やすく整形） ----
with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="SKU運用ランク")
    ws = writer.sheets["SKU運用ランク"]

    # 列幅自動調整
    col_widths = {"A":18,"B":45,"C":14,"D":14,"E":10,"F":12,"G":14,"H":12,"I":12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # ヘッダー色
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ランクセルに色を付ける（売上ランク=F列, 昨対ランク=H列）
    rank_colors = {
        "A": "C6EFCE",  # 緑
        "B": "FFEB9C",  # 黄
        "C": "FFC7CE",  # 赤薄
        "D": "FF0000",  # 赤
        "新商品": "BDD7EE",  # 青
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in [row[5], row[7]]:  # F=売上ランク, H=昨対ランク
            v = str(cell.value)
            if v in rank_colors:
                cell.fill = PatternFill("solid", fgColor=rank_colors[v])
            cell.alignment = Alignment(horizontal="center")

    # 昨対増減率に %フォーマット
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value is not None:
                cell.number_format = '#,##0.0"%"'

    # 売上金額 カンマ区切り
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

# CSV出力（バックアップ用）
df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

print("=== 完了 ===")
print(f"出力先: {OUTPUT_XLSX}")
print(f"  今年商品数: {len(df)}")
print(f"  去年データあり: {df['去年売上'].notna().sum()}")
print(f"  新商品: {(df['昨対ランク']=='新商品').sum()}")
print()
print("【売上ランク分布】")
print(df["売上ランク"].value_counts().sort_index().to_string())
print()
print("【昨対ランク分布】")
print(df["昨対ランク"].value_counts().sort_index().to_string())
print()
print("--- 上位10件プレビュー ---")
preview = df.head(10)[["品番","今年売上","去年売上","売上ランク","昨対増減率","昨対ランク"]]
print(preview.to_string(index=False))
