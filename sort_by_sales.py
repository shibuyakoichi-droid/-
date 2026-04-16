"""
生産管理ファイル 販売数順ソートスクリプト
==========================================
【使い方】
python sort_by_sales.py \
    --mgmt  production_mgmt_GoogleSheets_updated.xlsx \
    --sales item_order_0308171954.csv \
    --sets  set_0308173852_000017.csv \
    --output production_mgmt_sorted_by_sales.xlsx

【処理内容】
1. 販売CSVの商品数量をSKU単位（品番＋属性1名＋属性2名）で集計
2. セットCSVを使って販売数を構成品SKUに按分
   - 直接マッチ : 販売CSVの品番が生産管理に直接ある → そのまま集計
   - セット按分 : 販売CSVの品番がセットCSVの親商品コード →
                 構成品の構成数量比率で各SKUに按分して加算
3. 生産管理シートを販売数の多い順（降順）に並べ替え
4. 最終列に「今月販売数(CSV実績)」列を追加

【セット按分の詳細】
セットCSVには「親商品コード＋親属性１名」→「構成品商品コード＋構成品属性１名＋構成数量」
が定義されている。
例）tubutubu-babyleg / Aセット → キナリ×1, 杢オートミール×1, キナリつぶ×1
  → Aセットが498個売れた場合、各色に498個ずつ加算（構成数量×販売数）

【マッチしない品番について】
生産管理にもセット親にも該当しない品番（廃番・管理外品等）は無視してスキップします。
スクリプト実行時にログで表示されます。
"""

import argparse
import copy
import sys
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────
# 1. 販売CSVを読み込んでSKU別に集計
# ─────────────────────────────────────────
def load_sales(sales_path):
    df = pd.read_csv(sales_path, encoding='cp932', dtype=str)
    df['商品数量_数値'] = df['商品数量'].str.replace(',', '').pipe(pd.to_numeric, errors='coerce').fillna(0)
    df['属性１名'] = df['属性１名'].fillna('')
    df['属性２名'] = df['属性２名'].fillna('')
    # (品番, 属性1名, 属性2名) → 販売数量合計
    agg = df.groupby(['商品コード', '属性１名', '属性２名'])['商品数量_数値'].sum()
    return agg.to_dict()


# ─────────────────────────────────────────
# 2. セットCSVを読み込んで按分辞書を作成
# ─────────────────────────────────────────
def load_set_definition(sets_path):
    """
    戻り値: dict
      key  : (親商品コード, 親属性１名)
      value: list of (構成品商品コード, 構成品属性１名, 構成数量)
    """
    df = pd.read_csv(sets_path, encoding='cp932', dtype=str)
    df['構成品属性１名'] = df['構成品属性１名'].fillna('')
    df['親属性１名']   = df['親属性１名'].fillna('')
    df['構成数量']     = pd.to_numeric(df['構成数量'], errors='coerce').fillna(1)

    set_dict = defaultdict(list)
    for _, row in df.iterrows():
        parent_key = (row['親商品コード'], row['親属性１名'])
        set_dict[parent_key].append((
            row['構成品商品コード'],
            row['構成品属性１名'],
            int(row['構成数量'])
        ))
    return set_dict


# ─────────────────────────────────────────
# 3. 生産管理の全SKUに販売数を付与
# ─────────────────────────────────────────
def calc_sales_per_sku(sales_raw, set_dict, mgmt_codes):
    """
    sales_raw : load_sales()の戻り値
    set_dict  : load_set_definition()の戻り値
    mgmt_codes: 生産管理に存在する品番のset

    戻り値: dict  key=(品番, 属性1名, 属性2名)  value=販売数(float)
    """
    result = defaultdict(float)

    for (hinban, attr1, attr2), qty in sales_raw.items():
        if qty == 0:
            continue

        # ── ケース②を先に評価: セット按分（親商品コード＋親属性１名で照合）
        # ※ 販売CSVのattr2は'F'等が入ることがあるが、セット定義は2キー(品番, 属性1名)なのでattr2は無視
        # ※ tubutubu-babylegのように「品番が生産管理にも存在するが、
        #    販売CSVではAセット/Bセットという属性名で来る」ケースがあるため、
        #    セット定義照合を直接マッチより先に行う
        parent_key = (hinban, attr1)
        if parent_key in set_dict:
            components = set_dict[parent_key]
            for comp_code, comp_attr1, comp_qty in components:
                result[(comp_code, comp_attr1, '')] += qty * comp_qty
            continue

        # ── ケース①: 品番が生産管理に直接ある → そのままSKUに加算
        if hinban in mgmt_codes:
            result[(hinban, attr1, attr2)] += qty
            continue

        # ── ケース③: どこにもない → スキップ（ログ出力）
        print(f'  [スキップ] {hinban} / {attr1} / {attr2}  販売数={qty:.0f}')

    return result


# ─────────────────────────────────────────
# 4. xlsxを読み込んでソート＆新列追加して保存
# ─────────────────────────────────────────
def read_cell_data(cell):
    return {
        'value':         cell.value,
        'font':          copy.copy(cell.font),
        'fill':          copy.copy(cell.fill),
        'alignment':     copy.copy(cell.alignment),
        'border':        copy.copy(cell.border),
        'number_format': cell.number_format,
    }


def process_xlsx(mgmt_path, sales_sku, output_path):
    wb = load_workbook(mgmt_path)
    ws = wb['生産管理_2026年2月']

    HEADER_ROW = 3
    DATA_START  = 4
    MAX_COL     = ws.max_column

    # 全データ行を読み込む
    print('シートデータを読み込み中...')
    all_rows = []
    for row in range(DATA_START, ws.max_row + 1):
        hinban = ws.cell(row, 1).value
        if hinban is None:
            continue
        attr1 = ws.cell(row, 4).value or ''
        attr2 = ws.cell(row, 6).value or ''
        qty = sales_sku.get((str(hinban), str(attr1), str(attr2)), None)
        if qty is None:
            qty = sales_sku.get((str(hinban), str(attr1), ''), 0.0)
        all_rows.append({
            'hinban':    hinban,
            'attr1':     attr1,
            'attr2':     attr2,
            'sales_qty': qty,
            'cells':     [read_cell_data(ws.cell(row, c)) for c in range(1, MAX_COL + 1)]
        })

    print(f'  読み込み: {len(all_rows)}行  マッチ: {sum(1 for r in all_rows if r["sales_qty"]>0)}行')

    # 販売数降順ソート
    all_rows.sort(key=lambda x: x['sales_qty'], reverse=True)

    # 新列ヘッダー
    NEW_COL = MAX_COL + 1
    hc = ws.cell(HEADER_ROW, NEW_COL)
    hc.value         = '今月販売数\n(CSV実績)'
    hc.font          = Font(bold=True, name='Arial', size=9)
    hc.fill          = PatternFill('solid', start_color='D9E1F2')
    hc.alignment     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # データ書き戻し
    print('ソート済みデータを書き込み中...')
    for i, row_data in enumerate(all_rows):
        excel_row = DATA_START + i
        for col_idx, cd in enumerate(row_data['cells'], start=1):
            c = ws.cell(excel_row, col_idx)
            c.value         = cd['value']
            c.font          = cd['font']
            c.fill          = cd['fill']
            c.alignment     = cd['alignment']
            c.border        = cd['border']
            c.number_format = cd['number_format']
        nc = ws.cell(excel_row, NEW_COL)
        nc.value     = int(round(row_data['sales_qty']))
        nc.font      = Font(name='Arial', size=9)
        nc.alignment = Alignment(horizontal='right', vertical='center')
        nc.fill      = PatternFill('solid', start_color='F2F2F2' if row_data['sales_qty'] == 0 else 'E2EFDA')

    ws.column_dimensions[get_column_letter(NEW_COL)].width = 13

    wb.save(output_path)
    print(f'保存完了: {output_path}')


# ─────────────────────────────────────────
# main
# ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='生産管理ファイルを販売数順にソートする')
    parser.add_argument('--mgmt',   required=True, help='生産管理xlsx')
    parser.add_argument('--sales',  required=True, help='販売実績CSV')
    parser.add_argument('--sets',   required=True, help='セット定義CSV')
    parser.add_argument('--output', required=True, help='出力xlsx')
    args = parser.parse_args()

    print('=== 販売CSVを読み込み ===')
    sales_raw = load_sales(args.sales)
    print(f'  販売SKU数: {len(sales_raw)}')

    print('\n=== セット定義CSVを読み込み ===')
    set_dict = load_set_definition(args.sets)
    print(f'  セット親SKU数: {len(set_dict)}')

    print('\n=== 生産管理の品番一覧を取得 ===')
    wb_tmp = load_workbook(args.mgmt)
    ws_tmp = wb_tmp['生産管理_2026年2月']
    mgmt_codes = set()
    for row in range(4, ws_tmp.max_row + 1):
        h = ws_tmp.cell(row, 1).value
        if h:
            mgmt_codes.add(str(h))
    print(f'  生産管理品番数: {len(mgmt_codes)}')

    print('\n=== SKU別販売数を計算（セット按分込み） ===')
    print('  [スキップ]はどこにも紐づかない品番（廃番・管理外品）:')
    sales_sku = calc_sales_per_sku(sales_raw, set_dict, mgmt_codes)
    print(f'  販売数が付いたSKU数: {len(sales_sku)}')

    print('\n=== xlsxを更新 ===')
    process_xlsx(args.mgmt, sales_sku, args.output)

    print('\n完了！')


if __name__ == '__main__':
    main()
